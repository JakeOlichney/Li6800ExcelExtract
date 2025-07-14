# This code is designed to take multiple excel files from the Licor LI6800
# and combine them into one excel file with a grouping variable as the LI6800 file names.
# This works best if the title of each file is the same as the grouping variable.
# For example, my file for red oak m3 is set as 2025-05-19-2248_rom3_resp.xlsx
# For red oak g2 2025-05-19-2253_rog2_resp.xlsx
# This code would combine them with a corresponding column that includes the title of the file.
# From there you can use excel to seperate the column by the _ deliniator.
# This sets up the file for use in the fitacis function of plantecophys R package
# https://www.rdocumentation.org/packages/plantecophys/versions/1.4-6/topics/fitacis


#Libraries. set up python and run in terminal pip install pandas, openpyxl and pywin32
import os
import pandas as pd
from openpyxl import load_workbook
import win32com.client

# Configuration. Set source_folder as folder containing all licor excel files.
# Do not include any empty files. They will result in an error that will prevent output.
# All files must have the columns in selected_cols = [] below.
# This section sets the selected columns. If you want to add more, copy from the licor file
# and merge the column (set up in 3 rows), with a space between each section. Note examples below.
# This section also sets the amount of rows to extract. If you have fewer than 27, then it will still
# extract all rows, just with blank values in the resulting excel file.
# headers_start sets the start for each column header.
# data_start sets the start for data values.
source_folder = 'C:/Users/jacob/OneDrive/Documents/American Chestnut/2025_licor_meas'
output_file = 'combined_data.xlsx'
selected_cols = [
    'GasEx A µmol m⁻² s⁻¹',
    'GasEx Ci µmol mol⁻¹',
    'LeafQ Qin µmol m⁻² s⁻¹',
    'Meas Tleaf °C'
]
rows_to_extract = 27
headers_start = 14
data_start = 17

# Step 1: Force Recalculation with Excel COM (Windows only)
# Licor excel files use formulas to calculate values.
# Each file will need to be calculated to have values for output. Otherwise the resulting file will contain 0's
def recalculate_excel_formulas(folder):
    excel_files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    for file in excel_files:
        file_path = os.path.join(folder, file)
        print(f"Recalculating: {file}")
        try:
            wb = excel.Workbooks.Open(file_path)
            wb.RefreshAll()
            excel.CalculateFullRebuild()
            wb.Save()
            wb.Close(SaveChanges=True)
        except Exception as e:
            print(f"Skipped {file} due to error: {e}")

    excel.Quit()
    print("All files recalculated and saved.")

# Step 2: Extract header and data values from Excel
# Using settings in the configuration above.
def extract_values_from_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = []
    for col in ws.iter_cols(min_row=headers_start, max_row=headers_start+2, min_col=1, values_only=True):
        combined = ' '.join([str(cell) if cell else '' for cell in col]).strip()
        combined = ' '.join(combined.split())  # Normalize spacing
        headers.append(combined)

    data = []
    for i in range(data_start, data_start + rows_to_extract):
        row = [ws.cell(row=i, column=j+1).value for j in range(len(headers))]
        data.append(row)

    return pd.DataFrame(data, columns=headers)

# Step 3: Build combined DataFrame
#set output as an excel file.
def build_combined_dataset():
    excel_files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx')]
    combined_data = []

    for file in excel_files:
        file_path = os.path.join(source_folder, file)
        print(f"Processing: {file}")

        try:
            df = extract_values_from_excel(file_path)
            df = df[selected_cols].copy()
            df['SourceFile'] = file
            combined_data.append(df)
        except Exception as e:
            print(f"Skipped {file}: {e}")

    if combined_data:
        final_df = pd.concat(combined_data, ignore_index=True)
        final_df.to_excel(output_file, index=False)
        print(f"Combined data saved to: {os.path.abspath(output_file)}")
    else:
        print("No valid data to combine.")

recalculate_excel_formulas(source_folder)
build_combined_dataset()

# Output dataset will be in same folder as python project folder.
